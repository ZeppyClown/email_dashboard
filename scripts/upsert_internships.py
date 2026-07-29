#!/usr/bin/env python3
"""Turn a JOB-BLAST internship spreadsheet into digest_items rows.

The xlsx attached to NTU's JOB-BLAST digest carries Summary, Remuneration,
Vacancies, Commences, Occupation(s), Industry and Contract type per listing.
Early digests kept only title/employer/deadline/link, which is why internship
cards opened to an empty detail panel. Those columns *are* the panel.

What the spreadsheet does not have: year of study, GPA, tech stack, duration.
Those live only on the CareerAxis listing page behind an NTU login — this
script never invents them, it links out instead.

    python3 scripts/upsert_internships.py ~/Downloads/joblist.xlsx --dry-run
    python3 scripts/upsert_internships.py ~/Downloads/joblist.xlsx
"""
from __future__ import annotations

import argparse
import sys
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from _common import (ConfigError, build_meta, careeraxis_id, classify, clean,  # noqa: E402
                     die, parse_date, stable_id, upsert)

# Spreadsheet header → our name. Matched case-insensitively on a normalised
# header, so "Contract type" and "Contract Type" both land.
COLUMNS = {
    "position": "title", "job title": "title", "title": "title",
    "employer": "employer", "company": "employer",
    "summary": "summary",
    "remuneration": "remuneration", "salary": "remuneration",
    "vacancies": "vacancies",
    "commences": "commences", "start date": "commences",
    "occupation": "occupation", "occupations": "occupation", "occupation(s)": "occupation",
    "industry": "industry",
    "contract type": "contract_type",
    "country": "country",
    "closes": "closes", "closing date": "closes", "deadline": "closes",
    "published": "published",
    "added": "added",
    "link": "link", "url": "url", "job url": "link",
}


def norm(header: Any) -> str:
    return str(header or "").strip().lower().rstrip("*: ")


def read_rows(path: Path) -> list[dict[str, Any]]:
    """Read the first sheet, tolerating a title banner above the header row."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    header_at = next(
        (i for i, row in enumerate(grid[:10])
         if sum(1 for c in row if norm(c) in COLUMNS) >= 3),
        None,
    )
    if header_at is None:
        die(f"no recognisable header row in {path.name}. "
            f"Expected columns like Position, Employer, Closes, Summary.")

    headers = [COLUMNS.get(norm(c)) for c in grid[header_at]]
    rows = []
    for raw in grid[header_at + 1:]:
        row = {h: raw[i] for i, h in enumerate(headers) if h and i < len(raw)}
        if clean(row.get("title")):
            rows.append(row)
    return rows


def build_item(row: dict[str, Any], digest_day: str) -> dict[str, Any] | None:
    title = clean(row.get("title"))
    if not title:
        return None

    link = clean(row.get("link")) or clean(row.get("url"))
    employer = clean(row.get("employer"))
    category, source = classify(link)

    try:
        item_id = stable_id(link, fallback=f"{title}|{employer or ''}")
    except ValueError:
        return None

    # `summary` is the one-line card face; `details` is the drawer. Repeating
    # the summary in details is what made cards feel empty, so details carries
    # the spreadsheet's prose plus an explicit pointer to what it cannot know.
    summary = clean(row.get("summary"))
    detail_parts = []
    if summary:
        detail_parts.append(summary)
    if link and careeraxis_id(link):
        detail_parts.append(
            "Year of study, GPA, tech stack and duration are not in the JOB-BLAST "
            f"spreadsheet — they appear only on the CareerAxis listing: {link}"
        )

    return {
        "id": item_id,
        "digest_date": digest_day,
        "category": category,
        "source": source,
        "title": f"{title} — {employer}" if employer else title,
        "summary": (summary[:180] if summary else None),
        "details": "\n\n".join(detail_parts) or None,
        "meta": build_meta([
            ("Position", row.get("title")),
            ("Employer", employer),
            ("Remuneration", row.get("remuneration")),
            ("Vacancies", row.get("vacancies")),
            ("Commences", row.get("commences")),
            ("Contract type", row.get("contract_type")),
            ("Occupation", row.get("occupation")),
            ("Industry", row.get("industry")),
            ("Country", row.get("country")),
            ("Published", parse_date(row.get("published")) or row.get("published")),
            ("Closes", parse_date(row.get("closes")) or row.get("closes")),
            ("CareerAxis ID", careeraxis_id(link)),
        ]),
        "sender": "JOB-BLAST@ntu.edu.sg",
        "link": link,
        "deadline_date": parse_date(row.get("closes")),
    }


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx", nargs="+", type=Path)
    ap.add_argument("--digest-date", default=date.today().isoformat())
    ap.add_argument("--dry-run", action="store_true",
                    help="parse and report, write nothing")
    a = ap.parse_args()

    items: dict[str, dict[str, Any]] = {}
    for path in a.xlsx:
        if not path.exists():
            die(f"{path} not found")
        rows = read_rows(path)
        built = [i for i in (build_item(r, a.digest_date) for r in rows) if i]
        print(f"{path.name}: {len(rows)} rows → {len(built)} items")
        for item in built:
            items[item["id"]] = item          # last write wins within a run

    by_cat: dict[str, int] = {}
    for i in items.values():
        by_cat[i["category"]] = by_cat.get(i["category"], 0) + 1
    print(f"\n{len(items)} unique · " + " · ".join(f"{k} {v}" for k, v in sorted(by_cat.items())))

    if a.dry_run:
        for i in list(items.values())[:3]:
            print(f"\n  {i['id']}\n  {i['title']}\n  meta: {list(i['meta'])}")
        print(f"\ndry run — nothing written")
        return 0

    try:
        n = upsert(list(items.values()))
    except ConfigError as e:
        die(str(e))
    print(f"upserted {n} (status and note preserved on existing cards)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
